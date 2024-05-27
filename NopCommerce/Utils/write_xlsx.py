from openpyxl import load_workbook


class XlsxWriter:
    @staticmethod
    def write_login_test_result(test_case_id, actual, pass_fail, comments):
        wb = load_workbook("../Test_Reports/TC_LoginTest_Report.xlsx")
        sheet = wb.active

        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_id:
                sheet.cell(row=i, column=6, value=actual)  # write to column 'F'
                sheet.cell(row=i, column=7, value=pass_fail)  # write to column 'G'
                sheet.cell(row=i, column=8, value=comments)  # write to column 'H'
                break
        wb.save("../Test_Reports/TC_LoginTest_Report.xlsx")

    @staticmethod
    def write_signup_test_result(test_case_id, actual, pass_fail, comments):
        wb = load_workbook("../Test_Reports/TC_SignupTest_Report.xlsx")
        sheet = wb.active
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_id:
                sheet.cell(row=i, column=8, value=actual)  # write to column 'H'
                sheet.cell(row=i, column=9, value=pass_fail)  # write to column 'I'
                sheet.cell(row=i, column=10, value=comments)  # write to column 'J'
                break
        wb.save("../Test_Reports/TC_SignupTest_Report.xlsx")

    @staticmethod
    def write_add_to_cart_test_result(test_case_id, actual, pass_fail, comments):
        wb = load_workbook("../Test_Reports/TC_AddToCartTest_Report.xlsx")
        sheet = wb.active

        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_id:
                sheet.cell(row=i, column=4, value=actual)  # write to column 'D'
                sheet.cell(row=i, column=5, value=pass_fail)  # write to column 'E'
                sheet.cell(row=i, column=6, value=comments)  # write to column 'F'
                break
        wb.save("../Test_Reports/TC_AddToCartTest_Report.xlsx")

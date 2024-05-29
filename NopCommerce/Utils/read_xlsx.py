from openpyxl import load_workbook
from login_data import LogInData
from signup_data import SignUpData
from Checkout_data import CheckoutData

class XlsxReader:
    @staticmethod
    def get_login_data():
        wb = load_workbook("../Test_Reports/TC_LoginTest_Report.xlsx")
        sheet = wb.active
        data = {}

        for i in range(2, sheet.max_row+1):
            user = LogInData(
                sheet.cell(i, 3).value,  # get email from row (2, 3 & 4) col 'C'
                sheet.cell(i, 4).value  # get password from row (2, 3 & 4) col 'D'
            )
            test_case_id = sheet.cell(i, 1).value  # get test case ID from column 'A'
            data[test_case_id] = user  # use the test case ID as the key
        return data

    @staticmethod
    def get_signup_data():
        wb = load_workbook("../Test_Reports/TC_SignupTest_Report.xlsx")
        sheet = wb.active
        data = {}

        for i in range(2, sheet.max_row+1):
            user = SignUpData(
                sheet.cell(i, 3).value,   # get first name from row (2 & 3) col 'C'
                sheet.cell(i, 4).value,   # get last name from row (2 & 3) col 'D'
                sheet.cell(i, 5).value,   # get email from row (2 & 3) col 'E'
                sheet.cell(i, 6).value    # get password from row (2 & 3) col 'F'
            )
            test_case_id = sheet.cell(i, 1).value  # get test case ID from column 'A'
            data[test_case_id] = user  # use the test case ID as the key
        return data

    @staticmethod
    def get_checkout_data():
        wb = load_workbook("../Test_Reports/TC_CheckOut_Report .xlsx")
        sheet = wb.active
        data = {}

        for i in range(2, sheet.max_row + 1):
            user = CheckoutData(
                sheet.cell(i, 3).value,     # get first name from row (2, 3) col 'C'
                sheet.cell(i, 4).value,     # get last name from row (2, 3) col 'D'
                sheet.cell(i, 5).value,     # get email from row (2, 3) col 'E'
                sheet.cell(i, 6).value,     # get password from row (2, 3) col 'F'
                sheet.cell(i, 7).value,     # get country from row (2, 3) col 'G'
                sheet.cell(i, 8).value,     # get ciy from row (2, 3) col 'H'
                sheet.cell(i, 9).value,     # get address 1 from row (2, 3) col 'I'
                sheet.cell(i, 10).value,    # get zip/postal code from row (2, 3) col 'J'
                sheet.cell(i, 11).value,    # get phone number from row (2, 3) col 'K'
                sheet.cell(i, 12).value,    # get shipping option from row (2, 3) col 'L'
                sheet.cell(i, 13).value,    # get payment option from row (2, 3) col 'M'
            )
            test_case_id = sheet.cell(i, 1).value  # get test case ID from column 'A'
            data[test_case_id] = user  # use the test case ID as the key
        return data

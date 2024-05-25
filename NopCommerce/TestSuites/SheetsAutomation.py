from openpyxl import load_workbook


class SheetsAutomation:
    def __init__(self, filename) -> None:
        self.fileName = filename
        self.workbook = load_workbook(filename)
        self.sheet = self.workbook.active

    def set_active(self, filename):
        self.fileName = filename
        self.workbook = load_workbook(filename)
        self.sheet = self.workbook.active

    def read_from_cell(self, row, col):
        return self.sheet.cell(row, col).value

    def write_in_cell(self, row, col, data):
        self.sheet.cell(row, col, value=data)

    def save(self):
        self.workbook.save(self.fileName)
